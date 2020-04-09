# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

# import xlsx
from astroquery.simbad import Simbad
import os.path
import numpy as np
import xlrd
import xlwt
from xlutils.copy import copy
from contextlib import closing
from urllib.request import Request
from urllib.request import urlopen
import warnings
import pandas as pd
import openpyxl
from astropy.table import Table
from astropy.table import MaskedColumn
from astropy import units as u
from astroquery import sha
from astropy.coordinates import SkyCoord
from astroquery import open_exoplanet_catalogue as oec
# from astroquery.open_exoplanet_catalogue import findvalue
from requests.exceptions import MissingSchema
import dill as pickle

import cascade
from cascade.exoplanet_tools import extract_exoplanet_data

__all__ = ['read_target_list', 'get_orbital_parameters',
           'RetrieveAORs', 'fovid_to_band',
           'write_orbit_parameters_to_file',
           'read_orbit_parameters_from_file']


def read_target_list(file_name, path=''):
    """
    ####################################
    # Readinput target list from xml file
    ####################################
    """
    wb_targets = openpyxl.load_workbook(filename=os.path.join(path, file_name))
    ws1 = wb_targets.active
    targets = []
    # here you iterate over the rows in the specific column
    for row in range(3, ws1.max_row+1):
        for column in "A":
            cell_name = "{}{}".format(column, row)
            targets.append(ws1[cell_name].value)

    targets_stripped = [i.rsplit('A b', 1)[0] for i in targets]
    targets_stripped = [i.rsplit('N b', 1)[0] for i in targets_stripped]
    targets_stripped = [i.rsplit(' ', 1)[0] for i in targets_stripped]

    return(wb_targets, targets, targets_stripped)


def get_orbital_parameters(archive_name, target_list, stripped_target_list,
                           search_radius=11.0*u.arcsec):
    """
    Input:
        name of archive containing exoplanet data
    target_list
        list containing names of planets
    stripped_target_list
        list containg stellar names arong which the planets orbit
    search_radius
        radius within the catalogue will be search for planet
        around its coordinates on the sky
    """
    _VALID_ARCHIVE = ['NASAEXOPLANETARCHIVE']

    if archive_name.strip() not in _VALID_ARCHIVE:
        raise KeyError("exoplenet archive name not recognized")

    ################################################
    # Query SIMBAD to check if targets can be found
    #################################################
    customSimbad = Simbad()
    SimbadResult = customSimbad.query_objects(stripped_target_list)
    # check if all targets have been found
    assert np.all(~SimbadResult['MAIN_ID'].data.mask), \
        "Not all targets could be found on SIMBAD"

    ct = cascade.exoplanet_tools.parse_database('NASAEXOPLANETARCHIVE',
                                                update=True)
    TT = []
    PERIOD = []
    ECC = []
    OMEGA = []
    INCL = []
    for i, (name, name_stripped) in enumerate(zip(target_list,
                                                  stripped_target_list)):
        try:
            dr = extract_exoplanet_data(ct, name, search_radius=search_radius)
            print('NASAEXOPLANETARCHIVE: ', name, dr[0]['NAME'].data.data)
            TT.append(np.ma.array(dr[0]['TT'][0], dtype=np.float64))
            PERIOD.append(np.ma.array(dr[0]['PER'][0], dtype=np.float64))
            ECC.append(np.ma.array(dr[0]['ECC'][0], dtype=np.float64))
            OMEGA.append(np.ma.array(dr[0]['OM'][0], dtype=np.float64))
            INCL.append(np.ma.array(dr[0]['I'][0], dtype=np.float64))
        except (KeyError, ValueError) as e:
            print("Planet {} not found in NASAEXOPLANETARCHIVE".format(name))
            print(e)

    TT = np.ma.array([i.data for i in TT], mask=[i.mask for i in TT])
    PERIOD = np.ma.array([i.data for i in PERIOD],
                         mask=[i.mask for i in PERIOD])
    ECC = np.ma.array([i.data for i in ECC], mask=[i.mask for i in ECC])
    OMEGA = np.ma.array([i.data for i in OMEGA], mask=[i.mask for i in OMEGA])
    INCL = np.ma.array([i.data for i in INCL], mask=[i.mask for i in INCL])

    ParTable = Table()
    ParTable["NAME"] = MaskedColumn(target_list)
    ParTable["SIMBADNAME"] = SimbadResult['MAIN_ID']
    ParTable["RA"] = SimbadResult['RA']
    ParTable["DEC"] = SimbadResult['DEC']
    ParTable["TT"] = MaskedColumn(TT)
    ParTable["PERIOD"] = MaskedColumn(PERIOD)
    ParTable["ECC"] = MaskedColumn(ECC)
    ParTable["INCL"] = MaskedColumn(INCL)
    ParTable["OMEGA"] = MaskedColumn(OMEGA)
    return ParTable


def write_orbit_parameters_to_file(ParTable, file_name, path=''):
    """
    Input:
    ParTable
        table containing oribital parameters of planetary systems
    """

    ##################################
    # create xls file to store resutls
    ###################################
    alignWrap = xlwt.Alignment()
    alignWrap.horz = xlwt.Alignment.HORZ_CENTER
    alignWrap.vert = xlwt.Alignment.VERT_CENTER
    alignWrap.wrap = xlwt.Alignment.WRAP_AT_RIGHT
    style1 = xlwt.easyxf('font: bold 1, color black;')
    style1.alignment = alignWrap
    style2 = xlwt.easyxf('font: bold 1, color blue;')
    style2.alignment = alignWrap

    wb_out = xlwt.Workbook()
    sheet_orbit_parameters = wb_out.add_sheet("Orbit Parameters")
    sheet_orbit_parameters.row(0).set_style(style2)
    sheet_orbit_parameters.row(1).set_style(style2)
    sheet_orbit_parameters.write_merge(0, 1, 0, 0, 'TARGET', style2)
    sheet_orbit_parameters.col(0).width = 256 * 30
    sheet_orbit_parameters.write_merge(0, 1, 1, 1, 'SIMBADNAME', style2)
    sheet_orbit_parameters.col(1).width = 256 * 30
    sheet_orbit_parameters.write(0, 2, 'RA', style2)
    sheet_orbit_parameters.write(0, 3, 'DEC', style2)
    sheet_orbit_parameters.col(2).width = 256 * 20
    sheet_orbit_parameters.col(3).width = 256 * 20
    sheet_orbit_parameters.write(0, 4, 'TT', style2)
    sheet_orbit_parameters.col(4).width = 256 * 20
    sheet_orbit_parameters.write(0, 5, 'PERIOD', style2)
    sheet_orbit_parameters.write(0, 6, 'ECC', style2)
    sheet_orbit_parameters.write(0, 7, 'INCL', style2)
    sheet_orbit_parameters.write(0, 8, 'OMEGA', style2)
    sheet_orbit_parameters.write(1, 2, '[hourangle]', style2)
    sheet_orbit_parameters.write(1, 3, '[degrees]', style2)
    sheet_orbit_parameters.write(1, 4, '[BJD]', style2)
    sheet_orbit_parameters.write(1, 5, '[days]', style2)
    sheet_orbit_parameters.write(1, 6, '', style2)
    sheet_orbit_parameters.write(1, 7, '[degree]', style2)
    sheet_orbit_parameters.write(1, 8, '[degree]', style2)
    for ic, cn in enumerate(ParTable.keys()):
        for ir, rval in enumerate(ParTable[cn]):
            if isinstance(rval, bytes):
                sheet_orbit_parameters.write(ir+2, ic, str(rval.decode()),
                                             style1)
            else:
                sheet_orbit_parameters.write(ir+2, ic, str(rval), style1)

    wb_out.save(os.path.join(path, file_name))
    return


def read_orbit_parameters_from_file(file_name, path=''):
    """
    """
    df = pd.read_excel(os.path.join(path, file_name), skiprows=[1],
                       na_values=['--', '', -1.0],
                       dtype={'NAME': np.str, 'SIMBADNAME': np.str,
                                   'RA': np.str, 'DEC': np.str,
                                   'TT': np.float64,
                                   'PERIOD': np.float64, 'ECC': np.float64,
                                   'INCL': np.float64, 'OMEGA': np.float64})
    ParTable = Table(masked=True).from_pandas(df)

    return ParTable


def RetrieveAORs(planet_name_or_coord, search_radius):
    """
    Searches SSC observing logs and returns AOR list

    Example:
        df = RetrieveAORs(' 55  CnC e ', 10.0)
        df = RetrieveAORs('08 52 35.8109,+28 19 50.951', 10.0)

    Note, best search on position as names sometime fail.
    Note2: No real error handling implemented yet if search fails
    """
    DEFAULT_TIMEOUT = 30.0
    BROWSER_MASQUERADE = "Mozilla/5.0 [en]"
    URL = "https://heasarc.gsfc.nasa.gov/db-perl/W3Browse/w3table.pl"

    Params = (("Action=Start+Search"
               "&Entry={}"
               "&Radius={}&Radius_unit=arcsec"
               "&ResultMax=0"
               "&displaymode=PureTextDisplay"
               "&tablehead=name=heasarc_spitzmastr"
               "&Fields=All"
               "&Coordinates=J2000"
               "&CheckCaches/GRB/SIMBAD+Sesame/NED").
              format("+".join(planet_name_or_coord.split()), search_radius))

    req = Request(URL, Params.encode())
    req.add_header('User-agent', BROWSER_MASQUERADE)
    with closing(urlopen(req, timeout=DEFAULT_TIMEOUT)) as response:
        htmlReceived = response.read().decode('utf-8')

    ListReceived = [x.split('|') for x in htmlReceived.rstrip().split('\n')]
    if 'No matches for:' in ListReceived[2]:
        warnings.warn('No observations found for: {}'.
                      format(planet_name_or_coord))
        return False
    AORInfoTable = pd.DataFrame.from_records(ListReceived[4:])
    AORInfoTable.columns = AORInfoTable.iloc[0]
    AORInfoTable = AORInfoTable.reindex(AORInfoTable.index.drop(0))
    AORInfoTable.columns = \
        AORInfoTable.columns.str.strip().\
        str.lower().str.replace(' ', '_').\
        str.replace('(', '').str.replace(')', '')

    return AORInfoTable


def fovid_to_band(fovid_list):
    FOVID_DICT = {'IRAC_Center_of_3.6umSub-array': ['IRAC 3.6um'],
                  'IRAC_Center_of_4.5umSub-Array': ['IRAC 4.5um'],
                  'IRAC_Center_of_3.6&5.8umArray': ['IRAC 3.6um','IRAC 5.8um'],
                  'IRAC_Center_of_4.5&8.0umArray': ['IRAC 4.5um','IRAC 8.0um'],
                  'IRAC_Center_of_5.8umSub-array': ['IRAC 5.8um'],
                  'IRAC_Center_of_8.0umSub-Array': ['IRAC 8.0um'],
                  'IRS_Short-Lo_1st_Order_1st_Position': ['IRS SL'],
                  'IRS_Short-Lo_1st_Order_2nd_Position': ['IRS SL'],
                  'IRS_Short-Lo_2nd_Order_1st_Position': ['IRS SL'],
                  'IRS_Short-Lo_2nd_Order_2nd_Position': ['IRS SL'],
                  'IRS_Short-Lo_1st_Order_Center_Position': ['IRS SL'],
                  'IRS_Short-Lo_2nd_Order_Center_Position': ['IRS SL'],
                  'IRS_Long-Lo_1st_Order_Center_Position': ['IRS LL'],
                  'IRS_Long-Lo_2nd_Order_Center_Position': ['IRS LL'],
                  'IRS_Blue_Peak-Up_FOV_Center': ['IRS PU Blue'],
                  'IRS_Red_Peak-Up_FOV_Center': ['IRS PU Red'],
                  'MIPS_24um_small_FOV1': ['MIPS 24um'],
                  'MIPS_24um_small_FOV2': ['MIPS 24um'],
                  'MIPS_70um_default_small_FOV1': ['MIPS 70um'],
                  'MIPS_70um_default_small_FOV2': ['MIPS 70um']}
    band = []
    for fovid in fovid_list:
        band += FOVID_DICT[fovid.decode().strip()]
    band = np.unique(band).tolist()
    return band
